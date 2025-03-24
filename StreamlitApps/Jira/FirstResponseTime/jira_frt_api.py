import requests
import datetime
import pandas as pd
import streamlit as st
from decouple import config
from openpyxl import Workbook
from pathlib import Path
from typing import Dict, List, Optional


class JiraAPI:
    """
    Class to interact with the Jira API and retrieve first response time data.

    Attributes:
        base_url (str): The base URL for the Jira API.
        token (str): The authentication token for Jira API.
        headers (Dict[str, str]): Headers used for authentication in API requests.
    """

    def __init__(self) -> None:
        """
        Initializes the JiraAPI instance by retrieving API credentials from environment variables.
        """
        # Load Jira base URL and API token from environment variables
        self.base_url: str = config("JIRA_BASE_URL").rstrip("/")
        self.token: str = config("JIRA_TOKEN")

        # Set up headers for authentication when making API requests
        self.headers: Dict[str, str] = {
            "Accept": "application/json",
            "Authorization": f"Bearer {self.token}"
        }

    def get_first_response_time(self, project_id: int, days_interval: int) -> pd.DataFrame:
        """
        Fetches Jira issues for a given project within a specified time frame and calculates the first response time.

        Args:
            project_id (int): The Jira project ID.
            days_interval (int): The number of days from today to retrieve issues.

        Returns:
            pd.DataFrame: A DataFrame containing issue details and response time.
        """
        # Define Jira Query Language (JQL) query to fetch relevant issues
        jql_query: str = f"""
            project = {project_id} 
            AND issuetype = "Support Request" 
            AND created >= -{days_interval}d 
            ORDER BY created DESC
        """

        # Construct API endpoint URL
        url: str = f"{self.base_url}/rest/api/2/search"

        # Define request parameters including JQL query and required fields
        params: Dict[str, str] = {
            "jql": jql_query,
            "fields": "summary,created,reporter,assignee,priority,comment",
            "maxResults": "500"  # Limit results to 100 issues
        }

        # Make the API request
        response: requests.Response = requests.get(url, headers=self.headers, params=params)

        # Handle potential authentication failure
        if response.status_code == 401:
            raise Exception("Authentication failed: Check JIRA_TOKEN in .env")
        elif response.status_code != 200:
            raise Exception(f"Jira API Error {response.status_code}: {response.text}")

        # Parse JSON response
        issues_data: Dict = response.json()
        results: List[Dict[str, Optional[str]]] = []

        # Process each issue retrieved from Jira
        for issue in issues_data.get("issues", []):
            issue_key: str = issue["key"]  # Unique identifier for the issue
            summary: str = issue["fields"]["summary"]  # Issue summary/title
            created: str = issue["fields"]["created"]  # Creation timestamp
            reporter: str = issue["fields"]["reporter"]["displayName"]  # Name of the reporter
            assignee: str = issue["fields"].get("assignee", {}).get("displayName",
                                                                    "Unassigned")  # Assignee or "Unassigned"
            priority: str = issue["fields"]["priority"]["name"]  # Priority level

            # Extract comments and determine first response time
            comments: List[Dict[str, str]] = issue["fields"]["comment"]["comments"]
            if comments:
                first_response_time: str = comments[0]["created"]
                time_to_first_response: Optional[int] = (
                                                                datetime.datetime.strptime(first_response_time,
                                                                                           "%Y-%m-%dT%H:%M:%S.%f%z")
                                                                - datetime.datetime.strptime(created,
                                                                                             "%Y-%m-%dT%H:%M:%S.%f%z")
                                                        ).total_seconds() // 60  # Convert response time to minutes
            else:
                first_response_time = None
                time_to_first_response = None

            # Store extracted data in results list
            results.append({
                "Issue Key": issue_key,
                "Summary": summary,
                "Issue Created": created,
                "Reporter": reporter,
                "Assignee": assignee,
                "Priority": priority,
                "First Response": first_response_time,
                "Time to First Response (Minutes)": time_to_first_response
            })

        # Convert results list to DataFrame and return
        return pd.DataFrame(results)


class ExcelExporter:
    """
    Class to export pandas DataFrame data to an Excel file.

    Attributes:
        filename (str): The name of the output Excel file.
        workbook (Workbook): The OpenPyXL workbook instance.
    """

    def __init__(self, filename: str) -> None:
        """
        Initializes the ExcelExporter with the filename.
        """
        self.filename: str = filename
        self.workbook: Workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def export_to_excel(self, dataframe: pd.DataFrame, sheet_name: str) -> None:
        """
        Exports a given DataFrame to an Excel sheet.

        Args:
            dataframe (pd.DataFrame): The DataFrame to export.
            sheet_name (str): The name of the sheet in the Excel file.
        """
        if not dataframe.empty:
            worksheet = self.workbook.create_sheet(title=sheet_name)
            worksheet.append(list(dataframe.columns))
            for row in dataframe.itertuples(index=False):
                worksheet.append(row)

    def save_and_close(self) -> None:
        """
        Saves and closes the Excel file.
        """
        if len(self.workbook.sheetnames) == 0:
            raise Exception("No data available to export. Ensure at least one project has data.")
        self.workbook.save(self.filename)
        self.workbook.close()


class StreamlitApp:
    """
    Streamlit application to retrieve and display Jira first response time data.
    """
    project_mapping: Dict[str, int] = {
        'Asda': 10090,
        'Clarins': 13032,
        'CCHQwell': 15833,
        'Mayborn': 17431
    }

    def __init__(self) -> None:
        """
        Initializes the Streamlit application UI components.
        """
        st.title('Jira First Response Time')
        self.jira_api = JiraAPI()
        self.selected_projects = st.multiselect('Select Projects', list(self.project_mapping.keys()))
        self.days_interval = st.number_input('Days Interval', step=1, min_value=1)
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        self.excel_exporter = ExcelExporter(filename=f"JIRA_FRT_{timestamp}.xlsx")

    def run(self) -> None:
        """
        Executes the application logic, retrieving and displaying Jira data.
        """
        if not self.selected_projects:
            st.warning('Please select at least one project.')
        elif self.days_interval < 1:
            st.warning('Days interval must be at least 1.')
        else:
            if st.button('Get First Response Time'):
                with st.spinner('Retrieving data from JIRA...'):
                    try:
                        data_exported = False
                        for project_name in self.selected_projects:
                            project_id = self.project_mapping[project_name]
                            df = self.jira_api.get_first_response_time(project_id, self.days_interval)
                            if not df.empty:
                                sheet_name = project_name.replace(' ', '_')
                                self.excel_exporter.export_to_excel(df, sheet_name)
                                data_exported = True
                        if data_exported:
                            self.excel_exporter.save_and_close()
                            st.success('Data exported to Excel.')
                            self.display_download_button()
                        else:
                            st.warning("No data found for the selected projects.")
                    except Exception as e:
                        st.error(f'An error occurred: {e}')

    def display_download_button(self) -> None:
        if Path(self.excel_exporter.filename).is_file():
            st.download_button(
                label='Download Excel File with FRT',
                data=open(self.excel_exporter.filename, 'rb').read(),
                file_name=self.excel_exporter.filename,
                mime='application/octet-stream'
            )


def main() -> None:
    """
    Runs the Streamlit application.
    """
    app = StreamlitApp()
    app.run()


if __name__ == "__main__":
    main()
