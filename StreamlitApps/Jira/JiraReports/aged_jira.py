from decouple import config  # Library to read configuration from environment variables or .env file
from datetime import datetime, timedelta  # To handle date and time operations
from typing import List, Tuple, Dict  # To specify data types
from collections import defaultdict  # To create default dictionaries
import pandas as pd  # For data manipulation and analysis
from jira import JIRA, Issue  # JIRA client library
from openpyxl import load_workbook  # To read/write Excel files
from openpyxl.styles import Border, Side, Alignment, Font  # For styling Excel cells
from openpyxl.utils import get_column_letter  # To convert column numbers to letters
from openpyxl.comments import Comment  # To add comments to Excel cells
from projects_config import PROJECT_KEYS, PROJECT_NAMES  # Custom module to get project keys and names
from available_priorities import issue_priorities  # Custom module to get available issue priorities


class DateRangeCalculator:
    """
    A class used to calculate date ranges for JIRA issues based on their age.

    Methods
    -------
    get_date_ranges() -> Dict[str, Tuple[datetime, datetime]]:
        Static method to get predefined date ranges for categorizing JIRA issues.
    """

    @staticmethod
    def get_date_ranges() -> Dict[str, Tuple[datetime, datetime]]:
        """
        Get predefined date ranges for categorizing JIRA issues.

        Returns
        -------
        Dict[str, Tuple[datetime, datetime]]
            A dictionary with keys representing the category of issue age
            and values representing tuples of start and end dates for each range.
        """
        current_date = datetime.now()  # Get the current date and time
        start_of_current_month = current_date.replace(day=1)  # Get the first day of the current month

        # Define date ranges for categorizing issues based on their age
        date_ranges = {
            "not_closed_up_to_30": (
                start_of_current_month - timedelta(days=30),
                start_of_current_month - timedelta(days=1),
            ),
            "not_closed_between_31_60": (
                start_of_current_month - timedelta(days=60),
                start_of_current_month - timedelta(days=31)),
            "not_closed_between_61_90": (
                start_of_current_month - timedelta(days=90),
                start_of_current_month - timedelta(days=61)),
            "not_closed_more_than_90": (start_of_current_month -
                                        timedelta(days=90), None)
        }

        return date_ranges


class JQLQueryBuilder:
    """
    A class to build JIRA Query Language (JQL) queries for fetching JIRA issues.

    Attributes
    ----------
    project_keys : Tuple[str, ...]
        A tuple of project keys for which the JQL queries will be constructed.

    Methods
    -------
    construct_jql_queries() -> Dict[str, Dict[str, str]]:
        Construct JQL queries for different types of JIRA issues.
    """

    def __init__(self, projects_keys: Tuple[str, ...]):
        """
        Initialize the JQLQueryBuilder with project keys.

        Parameters
        ----------
        projects_keys : Tuple[str, ...]
            A tuple of project keys.
        """
        self.project_keys = projects_keys
        self.date_ranges = DateRangeCalculator.get_date_ranges()  # Get the date ranges

    def construct_jql_queries(self) -> Dict[str, Dict[str, str]]:
        """
        Construct JQL queries for different types of JIRA issues.

        Returns
        -------
        Dict[str, Dict[str, str]]
            A dictionary with keys representing the type of issue and values
            as dictionaries containing JQL queries for each issue category.
        """
        projects_query = ','.join(self.project_keys)  # Convert project keys to a comma-separated string

        # Build JQL queries for different types of issues
        return {
            "case_type_issue": self._build_case_type_issue_queries(
                projects_query),
            "case_type_not_issue": self._build_case_type_not_issue_queries(
                projects_query),
            "support_request": self._build_support_request_queries(
                projects_query),
        }

    def _build_case_type_issue_queries(self,
                                       projects_query: str) -> Dict[str, str]:
        """
        Build JQL queries for 'case_type_issue'.

        Parameters
        ----------
        projects_query : str
            A comma-separated string of project keys.

        Returns
        -------
        Dict[str, str]
            A dictionary containing JQL queries for 'case_type_issue'.
        """
        return {
            "created_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND "Case Type" = Issue AND '
                'created >= startOfMonth(-1) AND created < startOfMonth()'),
            "resolved_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND "Case Type" = Issue AND '
                'resolved >= startOfMonth(-1) AND resolved < startOfMonth()'),
            "closed_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND "Case Type" = Issue AND '
                'status CHANGED TO Closed DURING (startOfMonth(-1), endOfMonth(-1))'),
            "not_closed_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND "Case Type" = Issue AND '
                'status != Closed AND created >= startOfMonth(-1) AND created < startOfMonth()'
            ),
            "not_closed_between_31_60": self._build_jql_query(
                projects_query,
                f'issuetype = "Support Request" AND "Case Type" = Issue AND '
                f'status != Closed AND created >= '
                f'"{self.date_ranges["not_closed_between_31_60"][0].strftime("%Y-%m-%d")}" '
                f'AND created < '
                f'"{self.date_ranges["not_closed_between_31_60"][1].strftime("%Y-%m-%d")}"'
            ),
            "not_closed_between_61_90": self._build_jql_query(
                projects_query,
                f'issuetype = "Support Request" AND "Case Type" = Issue AND '
                f'status != Closed AND created >= '
                f'"{self.date_ranges["not_closed_between_61_90"][0].strftime("%Y-%m-%d")}" '
                f'AND created < '
                f'"{self.date_ranges["not_closed_between_61_90"][1].strftime("%Y-%m-%d")}"'
            ),
            "not_closed_more_than_90": self._build_jql_query(
                projects_query,
                f'issuetype = "Support Request" AND "Case Type" = Issue AND '
                f'status != Closed AND created < '
                f'"{self.date_ranges["not_closed_more_than_90"][0].strftime("%Y-%m-%d")}"'
            )
        }

    def _build_case_type_not_issue_queries(self,
                                           projects_query: str) -> Dict[str, str]:
        """
        Build JQL queries for 'case_type_not_issue'.

        Parameters
        ----------
        projects_query : str
            A comma-separated string of project keys.

        Returns
        -------
        Dict[str, str]
            A dictionary containing JQL queries for 'case_type_not_issue'.
        """
        return {
            "created_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND "Case Type" != Issue AND '
                'created >= startOfMonth(-1) AND created < startOfMonth()'),
            "resolved_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND "Case Type" != Issue AND '
                'resolved >= startOfMonth(-1) AND resolved < startOfMonth()'),
            "closed_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND "Case Type" != Issue AND '
                'status CHANGED TO Closed DURING (startOfMonth(-1), endOfMonth(-1))'),
            "not_closed_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND "Case Type" != Issue AND '
                'status != Closed AND created >= startOfMonth(-1) AND created < startOfMonth()'
            ),
            "not_closed_between_31_60": self._build_jql_query(
                projects_query,
                f'issuetype = "Support Request" AND "Case Type" != Issue AND '
                f'status != Closed AND created >= '
                f'"{self.date_ranges["not_closed_between_31_60"][0].strftime("%Y-%m-%d")}" '
                f'AND created < '
                f'"{self.date_ranges["not_closed_between_31_60"][1].strftime("%Y-%m-%d")}"'
            ),
            "not_closed_between_61_90": self._build_jql_query(
                projects_query,
                f'issuetype = "Support Request" AND "Case Type" != Issue AND '
                f'status != Closed AND created >= '
                f'"{self.date_ranges["not_closed_between_61_90"][0].strftime("%Y-%m-%d")}" '
                f'AND created < '
                f'"{self.date_ranges["not_closed_between_61_90"][1].strftime("%Y-%m-%d")}"'
            ),
            "not_closed_more_than_90": self._build_jql_query(
                projects_query,
                f'issuetype = "Support Request" AND "Case Type" != Issue AND '
                f'status != Closed AND created < '
                f'"{self.date_ranges["not_closed_more_than_90"][0].strftime("%Y-%m-%d")}"'
            )
        }

    def _build_support_request_queries(self,
                                       projects_query: str) -> Dict[str, str]:
        """
        Build JQL queries for 'support_request'.

        Parameters
        ----------
        projects_query : str
            A comma-separated string of project keys.

        Returns
        -------
        Dict[str, str]
            A dictionary containing JQL queries for 'support_request'.
        """
        return {
            "created_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND created >= startOfMonth(-1) '
                'AND created < startOfMonth()'),
            "resolved_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND resolved >= startOfMonth(-1) '
                'AND resolved < startOfMonth()'),
            "closed_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND status CHANGED TO Closed '
                'DURING (startOfMonth(-1), endOfMonth(-1))'),
            "not_closed_last_month": self._build_jql_query(
                projects_query,
                'issuetype = "Support Request" AND status != Closed AND '
                'created >= startOfMonth(-1) AND created < startOfMonth()'
            ),
            "not_closed_between_31_60": self._build_jql_query(
                projects_query,
                f'issuetype = "Support Request" AND status != Closed AND '
                f'created >= "{self.date_ranges["not_closed_between_31_60"][0].strftime("%Y-%m-%d")}" '
                f'AND created < '
                f'"{self.date_ranges["not_closed_between_31_60"][1].strftime("%Y-%m-%d")}"'
            ),
            "not_closed_between_61_90": self._build_jql_query(
                projects_query,
                f'issuetype = "Support Request" AND status != Closed AND '
                f'created >= "{self.date_ranges["not_closed_between_61_90"][0].strftime("%Y-%m-%d")}" '
                f'AND created < '
                f'"{self.date_ranges["not_closed_between_61_90"][1].strftime("%Y-%m-%d")}"'
            ),
            "not_closed_more_than_90": self._build_jql_query(
                projects_query,
                f'issuetype = "Support Request" AND status != Closed AND '
                f'created < "{self.date_ranges["not_closed_more_than_90"][0].strftime("%Y-%m-%d")}"'
            )
        }

    @staticmethod
    def _build_jql_query(projects_query: str, condition: str) -> str:
        """
        Build a JQL query string.

        Parameters
        ----------
        projects_query : str
            A comma-separated string of project keys.
        condition : str
            A condition string for the JQL query.

        Returns
        -------
        str
            A complete JQL query string.
        """
        return f'project in ({projects_query}) AND {condition}'  # Construct the full JQL query


class JiraConnector:
    """
    A context manager class for connecting to JIRA.

    Attributes
    ----------
    jira_server : str
        The base URL of the JIRA server.
    jira_api_token : str
        The API token for authenticating with the JIRA server.
    jira : JIRA
        The JIRA connection object.

    Methods
    -------
    __enter__() -> JIRA:
        Establish a connection to the JIRA server.
    __exit__(exc_type, exc_val, exc_tb):
        Close the connection to the JIRA server.
    connect() -> JIRA:
        Connect to the JIRA server and return the connection object.
    """

    def __init__(self):
        """
        Initialize the JiraConnector with configuration parameters.
        """
        self.jira_server = config('JIRA_BASE_URL')  # Read JIRA base URL from configuration
        self.jira_api_token = config('JIRA_TOKEN')  # Read JIRA API token from configuration
        self.jira = None  # Initialize JIRA connection object as None

    def __enter__(self):
        """
        Establish a connection to the JIRA server.

        Returns
        -------
        JIRA
            The JIRA connection object.
        """
        self.jira = self.connect()  # Connect to JIRA
        return self.jira  # Return the JIRA connection object

    def __exit__(self, exc_type, exc_val, exc_tb):
        """
        Close the connection to the JIRA server.

        Parameters
        ----------
        exc_type : type
            The exception type.
        exc_val : Exception
            The exception instance.
        exc_tb : traceback
            The traceback object.
        """
        if self.jira:
            self.jira.close()  # Close the JIRA connection if it exists

    def connect(self) -> JIRA:
        """
        Connect to the JIRA server and return the connection object.

        Returns
        -------
        JIRA
            The JIRA connection object.

        Raises
        ------
        Exception
            If there is an error while connecting to the JIRA server.
        """
        try:
            print("Connecting to JIRA...")  # Log the connection attempt
            return JIRA(server=self.jira_server, token_auth=self.jira_api_token)  # Connect to JIRA
        except Exception as exc:
            print(f"Failed to connect to JIRA: {exc}")  # Log the error if connection fails
            raise  # Raise the exception


class IssueFetcher:
    """
    A class to fetch JIRA issues using JQL queries.

    Attributes
    ----------
    jira : JIRA
        The JIRA connection object.

    Methods
    -------
    fetch_issues_with_jql(jql_query: str) -> List[Issue]:
        Fetch issues from JIRA based on the provided JQL query.
    """

    def __init__(self, jira: JIRA):
        """
        Initialize the IssueFetcher with a JIRA connection.

        Parameters
        ----------
        jira : JIRA
            The JIRA connection object.
        """
        self.jira = jira  # Assign the JIRA connection object to an instance variable

    def fetch_issues_with_jql(self, jql_query: str) -> List[Issue]:
        """
        Fetch issues from JIRA based on the provided JQL query.

        Parameters
        ----------
        jql_query : str
            The JQL query string.

        Returns
        -------
        List[Issue]
            A list of JIRA issues that match the JQL query.

        Raises
        ------
        Exception
            If there is an error while fetching issues from JIRA.
        """
        all_issues_list = []  # Initialize a list to store all fetched issues
        start_at = 0  # Starting point for pagination
        max_results = 1000  # Maximum number of results to fetch per request

        while True:  # Loop to handle pagination
            try:
                jira_issues = self.jira.search_issues(jql_query,
                                                      startAt=start_at,
                                                      maxResults=max_results)  # Fetch issues from JIRA
                all_issues_list.extend(jira_issues)  # Add fetched issues to the list

                if len(jira_issues) < max_results:  # Check if all issues have been fetched
                    break  # Break the loop if there are no more issues to fetch

                start_at += max_results  # Increment the start point for the next batch
            except Exception as ex:
                print(f"Failed to fetch issues for JQL: {jql_query} "
                      f"with exception: {ex}")  # Log the exception if fetching fails
                break  # Break the loop in case of an error

        return all_issues_list  # Return the list of fetched issues


class IssueCounter:
    """
    A class to count JIRA issues by project and priority.

    Methods
    -------
    count_issues_by_project_and_priority(
        issues: List[Issue], projects: List[str], priorities: List[str]
    ) -> Dict[str, Dict[str, Dict[str, Tuple[int, List[Tuple[str, str]]]]]]:
        Count the issues by project and priority.
    """

    @staticmethod
    def count_issues_by_project_and_priority(
            issues: List[Issue], projects: List[str], priorities: List[str]
    ) -> Dict[str, Dict[str, Dict[str, Tuple[int, List[Tuple[str, str]]]]]]:
        """
        Count the issues by project and priority.

        Parameters
        ----------
        issues : List[Issue]
            A list of JIRA issues.
        projects : List[str]
            A list of project keys.
        priorities : List[str]
            A list of priority names.

        Returns
        -------
        Dict[str, Dict[str, Dict[str, Tuple[int, List[Tuple[str, str]]]]]]
            A dictionary with counts of issues by project and priority.
        """
        count_dict = defaultdict(lambda: defaultdict(lambda: (0, [])))  # Initialize a nested default dictionary

        for project in projects:  # Initialize the count dictionary for all projects and priorities
            for priority in priorities:
                count_dict[project][priority] = (0, [])

        for issue in issues:  # Iterate over each issue
            project = issue.fields.project.key  # Get the project key of the issue
            priority = (issue.fields.priority.name if issue.fields.priority
                        else "No Priority")  # Get the priority of the issue
            count, details = count_dict[project][priority]  # Get the current count and details
            details.append((issue.key, issue.fields.summary))  # Add issue key and summary to details
            count_dict[project][priority] = (count + 1, details)  # Update the count and details

        return count_dict  # Return the count dictionary


class ExcelSaver:
    """
    A class to save issue counts to an Excel file.

    Methods
    -------
    save_counts_to_excel(
        filename: str,
        data: Dict[str, Dict[str, Dict[str, Tuple[int, List[Tuple[str, str]]]]]],
        projects: List[str],
        priorities: List[str]
    ):
        Save the issue counts to an Excel file.
    apply_formatting_to_excel(
        filename: str,
        data: Dict[str, Dict[str, Dict[str, Tuple[int, List[Tuple[str, str]]]]]],
        projects: List[str],
        priorities: List[str]
    ):
        Apply formatting to the Excel file.
    """

    @staticmethod
    def save_counts_to_excel(
            filename: str,
            data: Dict[str, Dict[str, Dict[str, Tuple[int, List[Tuple[str, str]]]]]],
            projects: List[str],
            priorities: List[str]
    ):
        """
        Save the issue counts to an Excel file.

        Parameters
        ----------
        filename : str
            The name of the Excel file.
        data : Dict[str, Dict[str, Dict[str, Tuple[int, List[Tuple[str, str]]]]]]
            The data to be saved to the Excel file.
        projects : List[str]
            A list of project keys.
        priorities : List[str]
            A list of priority names.
        """
        with pd.ExcelWriter(filename) as writer:  # Create an Excel writer
            for sheet_name, counts in data.items():  # Iterate over each sheet (query type)
                rows = []  # Initialize rows for the sheet
                header = ["Project", "Priority"]  # Define the header
                for query_name in counts.keys():  # Add query names to the header
                    header.append(query_name)
                rows.append(header)  # Add the header to the rows

                overall_totals = {query_name: [0] * len(priorities) for query_name in counts.keys()}

                for project in projects:  # Iterate over each project
                    project_name = PROJECT_NAMES.get(project, project)  # Get the project name
                    project_totals = [0] * len(counts.keys())

                    for i, priority in enumerate(priorities):  # Iterate over each priority
                        row = [project_name if i == 0 else "", priority]  # Add project name and priority to the row
                        for j, query_name in enumerate(counts.keys()):  # Add issue counts for each query
                            count, details = counts[query_name].get(
                                project, {}).get(priority, (0, []))
                            row.append(count)
                            project_totals[j] += count
                            overall_totals[query_name][i] += count
                        rows.append(row)  # Add the row to the rows

                    # Add total row for the project
                    total_row = ["", "Total"]
                    for total in project_totals:
                        total_row.append(total)
                    rows.append(total_row)

                    # Add two empty rows for spacing
                    rows.append([""] * len(header))
                    rows.append([""] * len(header))

                # Add overall totals for all projects
                for i, priority in enumerate(priorities):
                    total_row = ["All Projects" if i == 0 else "", priority]
                    for query_name in counts.keys():
                        total_row.append(overall_totals[query_name][i])
                    rows.append(total_row)

                total_summary_row = ["", "Total"]
                for j, query_name in enumerate(counts.keys()):  # Sum totals for each query
                    total_summary_row.append(sum(overall_totals[query_name]))
                rows.append(total_summary_row)

                df = pd.DataFrame(rows[1:], columns=rows[0])  # Create a DataFrame from the rows
                df.to_excel(writer, sheet_name=sheet_name, index=False)  # Write the DataFrame to the Excel sheet

        # Now apply formatting and comments
        ExcelSaver.apply_formatting_to_excel(filename, data, projects, priorities)

    @staticmethod
    def apply_formatting_to_excel(
            filename: str,
            data: Dict[str, Dict[str, Dict[str, Tuple[int, List[Tuple[str, str]]]]]],
            projects: List[str],
            priorities: List[str]
    ):
        """
        Apply formatting to the Excel file.

        Parameters
        ----------
        filename : str
            The name of the Excel file.
        data : Dict[str, Dict[str, Dict[str, Tuple[int, List[Tuple[str, str]]]]]]
            The data to be saved to the Excel file.
        projects : List[str]
            A list of project keys.
        priorities : List[str]
            A list of priority names.
        """
        workbook = load_workbook(filename)  # Load the workbook
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))  # Define a thin border
        current_date = datetime.now().strftime("%Y-%m-%d")  # Get the current date

        for sheet in workbook.worksheets:  # Iterate over each worksheet
            sheet.insert_rows(1)  # Insert a row at the top
            header_cell = sheet.cell(row=1, column=1)  # Get the first cell
            header_cell.value = (f'Jira raised as {sheet.title} (executed on '
                                 f'{current_date})')  # Set the header value
            header_cell.font = Font(bold=True, size=16)  # Set the header font
            header_cell.alignment = Alignment(horizontal='center',
                                              vertical='center')  # Center align the header
            sheet.merge_cells(start_row=1, start_column=1, end_row=1,
                              end_column=8)  # Merge cells for the header

            sheet.insert_rows(2)  # Insert another row

            for row in sheet.iter_rows(min_row=3):  # Iterate over each row starting from the third row
                for cell in row:  # Iterate over each cell in the row
                    if cell.value is not None and cell.value != "":
                        cell.border = thin_border  # Apply the border if the cell is not empty
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='center')  # Center align the cell

            for col in sheet.iter_cols(min_row=3, max_row=sheet.max_row,
                                       min_col=1, max_col=sheet.max_column):  # Iterate over each column
                max_length = 0
                column = get_column_letter(col[0].column)  # Get the column letter
                for cell in col:  # Iterate over each cell in the column
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))  # Find the maximum length of cell values
                adjusted_width = (max_length + 2)  # Adjust the column width
                sheet.column_dimensions[column].width = adjusted_width  # Set the column width

            max_length_col_a = 0
            for cell in sheet['A'][2:]:  # Iterate over each cell in the first column
                if cell.value:
                    max_length_col_a = max(max_length_col_a,
                                           len(str(cell.value)))  # Find the maximum length of cell values
            adjusted_width_col_a = max_length_col_a + 2  # Adjust the column width
            sheet.column_dimensions['A'].width = adjusted_width_col_a  # Set the column width

            project_row_offset = 0
            for project_idx, project in enumerate(projects):  # Iterate over each project
                for priority_idx, priority in enumerate(priorities):  # Iterate over each priority
                    for query_idx, query_name in enumerate(
                            data[sheet.title].keys()):  # Iterate over each query
                        cell = sheet.cell(row=4 + project_row_offset + priority_idx,
                                          column=3 + query_idx)  # Get the cell
                        if cell.value:
                            _, details = data[sheet.title][query_name].get(
                                project, {}).get(priority, (0, []))  # Get the issue details
                            if details:
                                comment_text = "\n".join([f"{issue_key}: {summary}"
                                                          for issue_key, summary in details])  # Create the comment text
                                cell.comment = Comment(comment_text, "JIRA")  # Add the comment to the cell
                project_row_offset += len(priorities) + 3  # Offset to account for the project rows and spacing

            # Adding comments to the "All Projects" section
            all_projects_start_row = project_row_offset + 3
            for priority_idx, priority in enumerate(priorities):
                for query_idx, query_name in enumerate(data[sheet.title].keys()):
                    cell = sheet.cell(row=all_projects_start_row + priority_idx,
                                      column=3 + query_idx)  # Get the cell
                    if cell.value:
                        _, details = data[sheet.title][query_name].get(
                            "All Projects", {}).get(priority, (0, []))  # Get the issue details
                        if details:
                            comment_text = "\n".join([f"{issue_key}: {summary}"
                                                      for issue_key, summary in details])  # Create the comment text
                            cell.comment = Comment(comment_text, "JIRA")  # Add the comment to the cell

        workbook.save(filename)  # Save the workbook


def main() -> None:
    project_keys = PROJECT_KEYS  # Get the project keys
    jql_query_builder = JQLQueryBuilder(project_keys)  # Initialize the JQL query builder
    jql_queries_dict = jql_query_builder.construct_jql_queries()  # Construct the JQL queries

    # Initialize results dictionary to store issue counts for different query types
    results = {
        "case_type_issue": defaultdict(lambda: defaultdict(dict)),
        "case_type_not_issue": defaultdict(lambda: defaultdict(dict)),
        "support_request": defaultdict(lambda: defaultdict(dict))
    }

    # Use the context manager to establish a connection to JIRA
    with JiraConnector() as jira_connection:
        issue_fetcher = IssueFetcher(jira_connection)  # Initialize the issue fetcher

        for query_type, queries in jql_queries_dict.items():  # Iterate over each query type
            for current_query_name, query_string in queries.items():  # Iterate over each query
                print(query_string)  # Print the JQL query
                fetched_issues = issue_fetcher.fetch_issues_with_jql(
                    query_string)  # Fetch issues using the JQL query
                # Count the issues by project and priority
                issue_counts = IssueCounter.count_issues_by_project_and_priority(
                    fetched_issues, PROJECT_KEYS, issue_priorities)
                results[query_type][current_query_name] = issue_counts  # Store the counts in the results dictionary

    # Save the issue counts to an Excel file
    ExcelSaver.save_counts_to_excel(
        f'AgedJira_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        results,
        PROJECT_KEYS,
        issue_priorities
    )


if __name__ == "__main__":
    main()
