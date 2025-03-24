from pathlib import Path
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference
from created_issues import IssueReport
from resolved_issues import JiraIssueFetcher
from closed_issues import generate_jira_reports
from StreamlitApps.Jira.JiraReports.ConfigModules.projects_config import PROJECT_NAMES
from StreamlitApps.Jira.JiraReports.ConfigModules.my_logger import configure_logger

# Configure the logger
logger = configure_logger(script_name=Path(__file__).stem)


class IssueDataFetcher:
    """
    A class to fetch issue data from various sources.

    Methods
    -------
    fetch_created_issues()
        Fetches created issues report.

    fetch_resolved_issues()
        Fetches resolved issues report.

    fetch_closed_issues()
        Fetches closed issues report.
    """

    @staticmethod
    def fetch_created_issues():
        """
        Fetches created issues report.

        Returns
        -------
        Tuple[Dict[str, pd.DataFrame], pd.DataFrame]
            A tuple containing a dictionary of project-specific dataframes and a total dataframe
            with created issues data.
        """
        try:
            created_issue_report = IssueReport()  # Initialize the report generation for created issues
            project_dataframes_created, total_df_created = created_issue_report.generate_report()  # Generate the report
            logger.info("Successfully fetched created issues.")  # Log success
            return project_dataframes_created, total_df_created  # Return the fetched data
        except Exception as e:
            logger.error(f"Failed to execute the report generation for Created issues: {e}", exc_info=True)  # Log error
            return {}, pd.DataFrame()  # Return empty data on failure

    @staticmethod
    def fetch_resolved_issues():
        """
        Fetches resolved issues report.

        Returns
        -------
        dict
            A dictionary containing project-specific and total resolved issues data.
        """
        try:
            issue_fetcher = JiraIssueFetcher()  # Initialize the fetcher for resolved issues
            result = issue_fetcher.fetch_and_report_issues()  # Fetch and report the issues
            logger.info("Successfully fetched resolved issues.")  # Log success
            return result  # Return the fetched data
        except Exception as e:
            logger.error(f"An error occurred in the main execution: {e}", exc_info=True)  # Log error
            return {}  # Return empty data on failure

    @staticmethod
    def fetch_closed_issues():
        """
        Fetches closed issues report.

        Returns
        -------
        Tuple[Dict[str, pd.DataFrame], pd.DataFrame]
            A tuple containing a dictionary of project-specific dataframes and a total dataframe
            with closed issues data.
        """
        try:
            project_issues_dfs, total_issues_df = generate_jira_reports()  # Generate the report for closed issues
            if project_issues_dfs is not None and total_issues_df is not None:
                logger.info("Successfully fetched closed issues.")  # Log success
                return project_issues_dfs, total_issues_df  # Return the fetched data
            else:
                logger.warning("Failed to retrieve the reports.")  # Log warning
                return {}, pd.DataFrame()  # Return empty data on failure
        except Exception as e:
            logger.error(f"Failed to fetch closed issues: {e}", exc_info=True)  # Log error
            return {}, pd.DataFrame()  # Return empty data on failure


class DataFrameMerger:
    """
    A class to merge different dataframes into a unified format.

    Methods
    -------
    merge_dataframes(project_dfs_created, project_dfs_resolved, project_dfs_closed, total_df_created, total_df_resolved, total_df_closed)
        Merges created, resolved, and closed issues dataframes into a single unified dataframe.
    """

    @staticmethod
    def merge_dataframes(project_dfs_created, project_dfs_resolved, project_dfs_closed, total_df_created,
                         total_df_resolved, total_df_closed):
        """
        Merges created, resolved, and closed issues dataframes into a single unified dataframe.

        Parameters
        ----------
        project_dfs_created : dict
            Dictionary of dataframes with created issues data for each project.

        project_dfs_resolved : dict
            Dictionary of dataframes with resolved issues data for each project.

        project_dfs_closed : dict
            Dictionary of dataframes with closed issues data for each project.

        total_df_created : pd.DataFrame
            Dataframe with total created issues data.

        total_df_resolved : pd.DataFrame
            Dataframe with total resolved issues data.

        total_df_closed : pd.DataFrame
            Dataframe with total closed issues data.

        Returns
        -------
        Tuple[Dict[str, pd.DataFrame], pd.DataFrame]
            A tuple containing a dictionary of merged project-specific dataframes and a merged total dataframe.
        """
        try:
            merged_project_dfs = {}

            # Merge data for each project
            for project in set(project_dfs_created.keys()).union(project_dfs_resolved["projects"].keys()).union(
                    project_dfs_closed.keys()):
                created_df = project_dfs_created.get(project, pd.DataFrame(columns=["Month"]))
                resolved_df = project_dfs_resolved["projects"].get(project, pd.DataFrame(columns=["Month"]))
                closed_df = project_dfs_closed.get(project, pd.DataFrame(columns=["Month"]))

                # Reverse the order of the resolved dataframe if it is not empty
                if not resolved_df.empty:
                    resolved_df = resolved_df.sort_values(by="Month").reset_index(drop=True)
                    resolved_df.iloc[:, 1] = resolved_df.iloc[:, 1].iloc[::-1].values

                # Initialize merged dataframe with appropriate columns
                merged_df = pd.DataFrame(columns=["Month", "Created", "Resolved", "Closed"])
                merged_df["Month"] = created_df["Month"] if "Month" in created_df else (
                    resolved_df["Month"] if "Month" in resolved_df else closed_df["Month"])
                merged_df["Created"] = created_df.iloc[:, 1] if len(created_df.columns) > 1 else pd.NA
                merged_df["Resolved"] = resolved_df.iloc[:, 1] if len(resolved_df.columns) > 1 else pd.NA
                merged_df["Closed"] = closed_df.iloc[:, 1] if len(closed_df.columns) > 1 else pd.NA

                merged_project_dfs[project] = merged_df

            # Merge total dataframes
            total_merged_df = pd.DataFrame(columns=["Month", "Created", "Resolved", "Closed"])
            total_merged_df["Month"] = total_df_created["Month"] if "Month" in total_df_created else (
                total_df_resolved["Month"] if "Month" in total_df_resolved else total_df_closed["Month"])
            total_merged_df["Created"] = total_df_created.iloc[:, 1] if len(total_df_created.columns) > 1 else pd.NA
            if len(total_df_resolved.columns) > 1:
                total_df_resolved = total_df_resolved.sort_values(by="Month").reset_index(drop=True)
                total_df_resolved.iloc[:, 1] = total_df_resolved.iloc[:, 1].iloc[::-1].values
            total_merged_df["Resolved"] = total_df_resolved.iloc[:, 1] if len(total_df_resolved.columns) > 1 else pd.NA
            total_merged_df["Closed"] = total_df_closed.iloc[:, 1] if len(total_df_closed.columns) > 1 else pd.NA

            logger.info("Successfully merged dataframes.")  # Log success
            return merged_project_dfs, total_merged_df  # Return merged dataframes
        except Exception as e:
            logger.error(f"Error occurred while merging dataframes: {e}", exc_info=True)  # Log error
            return {}, pd.DataFrame()  # Return empty data on failure


class ExcelReportGenerator:
    """
    A class to generate Excel reports from merged dataframes.

    Attributes
    ----------
    filename : str
        The name of the Excel file to save the report.

    Methods
    -------
    save_to_excel(merged_project_dfs, total_merged_df)
        Saves the merged dataframes to an Excel file with charts.
    """

    def __init__(self, filename):
        """
        Constructs all the necessary attributes for the ExcelReportGenerator object.

        Parameters
        ----------
        filename : str
            The name of the Excel file to save the report.
        """
        self.filename = filename

    def save_to_excel(self, merged_project_dfs, total_merged_df):
        """
        Saves the merged dataframes to an Excel file with charts.

        Parameters
        ----------
        merged_project_dfs : dict
            Dictionary of merged project-specific dataframes.

        total_merged_df : pd.DataFrame
            Merged total dataframe.
        """
        try:
            with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
                execution_date = datetime.now().strftime('%Y-%m-%d')
                execution_month = (datetime.now().replace(day=1) - pd.DateOffset(months=1)).strftime('%B %Y')

                # Save the total dataframe to an Excel sheet
                self._save_total_dataframe(writer, total_merged_df, execution_date, execution_month)
                # Save each project's dataframe to separate Excel sheets
                self._save_project_dataframes(writer, merged_project_dfs, execution_date, execution_month)
            logger.info(f"Successfully saved report to {self.filename}.")  # Log success
        except Exception as e:
            logger.error(f"Failed to save report to Excel: {e}", exc_info=True)  # Log error

    def _save_total_dataframe(self, writer, total_merged_df, execution_date, execution_month):
        """
        Saves the total merged dataframe to an Excel sheet with a chart.

        Parameters
        ----------
        writer : pd.ExcelWriter
            The Excel writer object.

        total_merged_df : pd.DataFrame
            The total merged dataframe.

        execution_date : str
            The date of report execution.

        execution_month : str
            The month of the data in the report.
        """
        # Write total dataframe to Excel sheet
        total_merged_df.to_excel(writer, sheet_name='Total', index=False, startrow=1)
        worksheet = writer.sheets['Total']
        worksheet.merge_cells('A1:I1')
        worksheet['A1'] = f'Jira Monthly Report (Total), Executed on: {execution_date}, for month: {execution_month}'
        worksheet['A1'].font = Font(bold=True)  # Set font to bold
        self._create_chart(writer, 'Total', total_merged_df)  # Create chart for total data

    def _save_project_dataframes(self, writer, merged_project_dfs, execution_date, execution_month):
        """
        Saves the project-specific merged dataframes to Excel sheets with charts.

        Parameters
        ----------
        writer : pd.ExcelWriter
            The Excel writer object.

        merged_project_dfs : dict
            Dictionary of merged project-specific dataframes.

        execution_date : str
            The date of report execution.

        execution_month : str
            The month of the data in the report.
        """
        # Define the order of projects
        project_order = ['ASD', 'CLRG', 'CCHV', 'MAYB', 'DNY']

        for project in project_order:
            if project in merged_project_dfs:
                project_name = PROJECT_NAMES.get(project, project)  # Get the project name
                df = merged_project_dfs[project]  # Get the project's dataframe
                df.to_excel(writer, sheet_name=project, index=False, startrow=1)  # Write dataframe to Excel sheet
                worksheet = writer.sheets[project]
                worksheet.merge_cells('A1:I1')
                worksheet['A1'] = (f'Jira Monthly Report ({project_name}), '
                                   f'Executed on: {execution_date}, for month: {execution_month}')
                worksheet['A1'].font = Font(bold=True)  # Set font to bold
                self._create_chart(writer, project, df)  # Create chart for project data

    @staticmethod
    def _create_chart(writer, sheet_name, df):
        """
        Creates a line chart for the given dataframe and adds it to the Excel sheet.

        Parameters
        ----------
        writer : pd.ExcelWriter
            The Excel writer object.

        sheet_name : str
            The name of the sheet to add the chart to.

        df : pd.DataFrame
            The dataframe to create the chart from.
        """
        worksheet = writer.sheets[sheet_name]
        chart = LineChart()  # Initialize a line chart
        chart.title = sheet_name
        chart.style = 2
        chart.y_axis.title = 'Count'
        chart.x_axis.title = 'Month'
        chart.y_axis.majorGridlines = None  # Remove major gridlines

        data = Reference(worksheet, min_col=2, min_row=2, max_col=4, max_row=len(df) + 1)  # Data range
        categories = Reference(worksheet, min_col=1, min_row=2, max_row=len(df) + 1)  # Category (x-axis) range

        chart.add_data(data, titles_from_data=True)  # Add data to the chart
        chart.set_categories(categories)  # Set categories

        chart.anchor = 'A1'
        chart.width = 36  # Set chart width
        chart.height = 27  # Set chart height

        chart_sheet_name = f'{sheet_name}_Chart'  # Define chart sheet name
        writer.book.create_sheet(chart_sheet_name)  # Create a new sheet for the chart
        chart_sheet = writer.sheets[chart_sheet_name]
        chart_sheet.add_chart(chart, 'A1')  # Add the chart to the new sheet

        # Hide extra rows and columns to focus on the chart
        for row in range(30, chart_sheet.max_row + 1):
            chart_sheet.row_dimensions[row].hidden = True
        for col in range(24, chart_sheet.max_column + 1):
            chart_sheet.column_dimensions[col].hidden = True


class ReportGenerator:
    """
    A class to generate the final report by orchestrating data fetching, merging, and Excel generation.

    Attributes
    ----------
    data_fetcher : IssueDataFetcher
        An instance of IssueDataFetcher to fetch issue data.

    merger : DataFrameMerger
        An instance of DataFrameMerger to merge fetched data.

    excel_generator : ExcelReportGenerator
        An instance of ExcelReportGenerator to generate the Excel report.

    Methods
    -------
    generate_report()
        Generates the final report by fetching, merging, and saving data to an Excel file.
    """

    def __init__(self, data_fetcher, merger, excel_generator):
        """
        Constructs all the necessary attributes for the ReportGenerator object.

        Parameters
        ----------
        data_fetcher : IssueDataFetcher
            An instance of IssueDataFetcher to fetch issue data.

        merger : DataFrameMerger
            An instance of DataFrameMerger to merge fetched data.

        excel_generator : ExcelReportGenerator
            An instance of ExcelReportGenerator to generate the Excel report.
        """
        self.data_fetcher = data_fetcher
        self.merger = merger
        self.excel_generator = excel_generator

    def generate_report(self):
        """
        Generates the final report by fetching, merging, and saving data to an Excel file.
        """
        logger.info("Starting report generation process.")  # Log the start of the report generation process

        # Fetch data for created, resolved, and closed issues
        project_dfs_created, total_df_created = self.data_fetcher.fetch_created_issues()
        resolved_result = self.data_fetcher.fetch_resolved_issues()
        project_dfs_closed, total_df_closed = self.data_fetcher.fetch_closed_issues()

        # Merge the fetched data
        merged_project_dfs, total_merged_df = self.merger.merge_dataframes(
            project_dfs_created,
            resolved_result,
            project_dfs_closed,
            total_df_created,
            resolved_result["total_count_per_month"],
            total_df_closed
        )

        # Save the merged data to an Excel file
        self.excel_generator.save_to_excel(merged_project_dfs, total_merged_df)
        logger.info("Report generation process completed.")  # Log the completion of the report generation process


def main():
    """
    The main function to initiate the report generation process.
    """
    excel_filename = f'Monthly_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'  # Generate the filename

    data_fetcher: IssueDataFetcher = IssueDataFetcher()  # Initialize the data fetcher
    merger: DataFrameMerger = DataFrameMerger()  # Initialize the dataframe merger
    excel_generator: ExcelReportGenerator = ExcelReportGenerator(
        excel_filename)  # Initialize the Excel report generator with the filename
    report_generator: ReportGenerator = ReportGenerator(data_fetcher, merger,
                                                        excel_generator)  # Initialize the report generator

    report_generator.generate_report()  # Generate the report


if __name__ == "__main__":
    main()
