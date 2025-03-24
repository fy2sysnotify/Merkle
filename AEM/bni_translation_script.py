import sys
import pandas as pd
import openpyxl
from typing import List, Optional, Tuple
from dataclasses import dataclass


@dataclass
class TranslationData:
    """
    Represents translation data for a language.

    Attributes:
        word (str): The English word to translate.
        translations (List[Optional[str]]): Translations in multiple languages.
    """

    word: str
    translations: List[Optional[str]]


@dataclass
class ExcelTranslator:
    """
    Represents an Excel translator.

    Attributes:
        excel_file_path (str): The path to the Excel file to be processed.
        skiprows (int, optional): The number of rows to skip at the beginning of the Excel sheet (default is 6).
        columns (List[str], optional): List of language columns (default is None).
        translations (List[TranslationData], optional): List of translation data.
        english (List[str], optional): List of English words.
    """

    excel_file_path: str
    skiprows: int = 6
    columns: List[str] = None
    translations: List[TranslationData] = None
    english: List[str] = None

    def __post_init__(self) -> None:
        if self.columns is None:
            self.columns: List[str] = []
        if self.translations is None:
            self.translations: List[TranslationData] = []
        if self.english is None:
            self.english: List[str] = []

    def process_translations(self, df: pd.DataFrame, col_pos: int) -> Tuple[pd.DataFrame, bool]:
        """
        Processes translations for a given language column.

        Args:
            df (pd.DataFrame): The DataFrame containing Excel data.
            col_pos (int): The position of the language column to be processed.

        Returns:
            Tuple[pd.DataFrame, bool]: A tuple containing the modified DataFrame and a boolean indicating if changes were made.
        """

        lang = self.columns[col_pos]
        changed = False

        for index, row in df.iterrows():
            en_value = row['en']
            lang_value = row[lang]

            if not (pd.notna(en_value) and pd.notna(lang_value)):
                continue

            for current in self.translations:
                if not (current.word and current.translations[col_pos - 1]):
                    continue

                if current.word.lower() in en_value.lower() and en_value != lang_value:
                    print(f"Replacing '{current.word}' with '{current.translations[col_pos - 1]}'")
                    df.at[index, lang] = lang_value.replace(current.word, current.translations[col_pos - 1])
                    changed = True

        return df, changed

    @staticmethod
    def process_words(current: List[Optional[str]], value: str, pos: int) -> List[Optional[str]]:
        """
        Processes and splits words from a cell value.

        Args:
            current (List[Optional[str]]): The current list of words for a language.
            value (str): The cell value to be processed.
            pos (int): The position of the language column.

        Returns:
            List[Optional[str]]: The updated list of words for the language.
        """

        parts = value.split('\n')

        for part in parts:
            if len(part) <= 3:
                continue

            if len(current) <= pos:
                current.append(None)

            current[pos] = part.strip()

        return current

    def load_translations(self) -> None:
        """
        Loads translation data from the Excel file.
        """

        try:
            df = pd.read_excel(self.excel_file_path, header=None, skiprows=self.skiprows)

            for index, row in df.iterrows():
                english_bits = row[4]

                if not pd.notna(english_bits):
                    continue

                current = self.process_words([], english_bits, 0)

                for col in range(1, len(self.columns)):
                    cell = row[col + 5]

                    if not pd.notna(cell):
                        continue

                    self.process_words(current, cell, col)

                for word in current:
                    if word and word not in self.english:
                        translations = [None] * (len(self.columns) - 1)
                        self.translations.append(TranslationData(word, translations))
                        self.english.append(word)

            self.translations.sort(key=lambda x: x.word if x.word else '')
        except Exception as e:
            print(f'Error: {e}')

    def process_excel(self, temp_file_path: str, sheet_name: str) -> None:
        """
        Processes translations for all language columns.

        Args:
            temp_file_path (str): The path to the temporary Excel file.
            sheet_name (str): The name of the sheet to be saved in the temporary file.
        """

        try:
            df = pd.read_excel(self.excel_file_path, header=None, skiprows=self.skiprows)

            for col in range(1, len(self.columns)):
                lang = self.columns[col]

                if not lang:
                    continue

                print(f'Processing {lang} [{col}]')
                df_copy, changed = self.process_translations(df.copy(), col)

                if not changed:
                    continue

                self.save_changes(df_copy, temp_file_path, sheet_name)
        except Exception as e:
            print(f"Error: {e}")

    def save_changes(self, df_copy: pd.DataFrame, temp_file_path: str, sheet_name: str) -> None:
        """
        Saves changes to the original Excel file.

        Args:
            df_copy (pd.DataFrame): The modified DataFrame to be saved.
            temp_file_path (str): The path to the temporary Excel file.
            sheet_name (str): The name of the sheet to be saved in the temporary file.
        """

        try:
            existing_workbook = openpyxl.load_workbook(self.excel_file_path)
            new_workbook = openpyxl.Workbook()
            new_workbook.remove(new_workbook.active)

            for sheet_name in existing_workbook.sheetnames:
                existing_sheet = existing_workbook[sheet_name]
                new_sheet = new_workbook.create_sheet(title=sheet_name)

                for row in existing_sheet.iter_rows():
                    new_sheet.append([cell.value for cell in row])

            with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
                df_copy.to_excel(writer, header=False, index=False, startrow=6, sheet_name=sheet_name)

            temp_workbook = openpyxl.load_workbook(temp_file_path)
            temp_sheet = temp_workbook.active

            if sheet_name in new_workbook.sheetnames:
                new_workbook.remove(new_workbook[sheet_name])

            new_workbook.create_sheet(title=sheet_name)

            for row in temp_sheet.iter_rows():
                new_workbook[sheet_name].append([cell.value for cell in row])

            new_workbook.save(self.excel_file_path)
        except Exception as e:
            print(f'Error: {e}')


def main() -> None:
    """
    Main function to perform translation processing.
    """

    # Check if the script was called with the correct number of arguments
    if len(sys.argv) != 2:
        print('Usage: python script.py <excel_file_path>')
        return

    # Retrieve the Excel file path from the command line argument
    excel_file_path: str = sys.argv[1]

    # Define a list of language columns for processing
    columns: List[str] = ['en', 'uk', 'fr', 'de', 'nl', 'el', 'da', 'no', 'fi', 'sv', 'is', 'lt', 'et', 'lv',
                          'hr', 'mk', 'sr', 'sq', 'bg', 'cs', 'hu', 'sk', 'sl', 'ro', 'pl', 'es', 'pt', 'it']

    # Create an instance of the ExcelTranslator class with the specified arguments
    excel_translator: ExcelTranslator = ExcelTranslator(excel_file_path, columns=columns)

    # Load translation data from the Excel file
    excel_translator.load_translations()

    # Specify the path for the temporary Excel file and the sheet name
    temp_file_path: str = 'temp.xlsx'
    sheet_name: str = 'YourSheetName'

    # Process translations and save changes to the Excel file
    excel_translator.process_excel(temp_file_path, sheet_name)


if __name__ == '__main__':
    main()
